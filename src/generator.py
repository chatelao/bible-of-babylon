import io
from pathlib import Path
from typing import List, Dict, Any
from jinja2 import Environment, FileSystemLoader, select_autoescape
import openpyxl
from openpyxl.cell.rich_text import CellRichText, InlineFont, TextBlock
from openpyxl.styles import Alignment, Font
from pygments import lex
from pygments.lexers import get_lexer_by_name
from pygments.util import ClassNotFound
from pygments.token import Token
from .models import (
    Pattern, Program, Instance, AnonymousInstance, Block, ListLiteral, Identifier,
    CallInstruction, AssignInstruction, ReturnInstruction, RawInstruction
)

def format_table_cell(value: Any, is_code: bool = False, lexer: str = None) -> str:
    str_value = str(value)
    if str_value != "N/A":
        str_value = str_value.rstrip()

    if is_code and str_value != "N/A":
        marker = f".. code-block:: {lexer}" if lexer else "::"
        return f"{marker}\n\n" + "\n".join(f"    {line}" for line in str_value.splitlines())

    if isinstance(value, str) and "\n" in str_value:
        lines = str_value.splitlines()
        return "\n".join(f"| {line}" if line.strip() else "|" for line in lines)
    return str_value

def format_value(value) -> str:
    if isinstance(value, str):
        return value
    if isinstance(value, (int, float)):
        return str(value)
    if isinstance(value, Identifier):
        return value.name
    if isinstance(value, ListLiteral):
        elements = [format_value(e) for e in value.elements]
        return f"[{', '.join(elements)}]"
    if isinstance(value, AnonymousInstance):
        assignments = [f"{a.name}={format_value(a.value)}" for a in value.assignments]
        return f"{value.pattern_name}({', '.join(assignments)})"
    if isinstance(value, Block):
        instructions = [format_value(i) for i in value.instructions]
        return f"{{ {'; '.join(instructions)} }}"
    if isinstance(value, CallInstruction):
        args = [format_value(a) for a in value.arguments]
        return f"call {value.name}({', '.join(args)})"
    if isinstance(value, AssignInstruction):
        return f"assign {value.target} = {format_value(value.value)}"
    if isinstance(value, ReturnInstruction):
        return f"return {format_value(value.value)}"
    if isinstance(value, RawInstruction):
        return f'raw "{value.snippet}"'
    return str(value)

class CodeGenerator:
    # Centralized lists of supported languages and formats
    PROGRAMMING_LANGUAGES = [
        "SQL", "C", "Cpp", "XQuery", "Java", "Rust", "Erlang", "Lisp", "Bash", "Cmd",
        "PowerShell", "Python", "PHP", "CSS", "CUDA", "x86 Assembler", "RISC-V Assembler", "Prolog",
        "X86", "Riscv", "Java Bytecode", "OCaml", "ARM AArch64 Assembler", "Arm", "Camel",
        "Go", "Haskell", "TypeScript", "Tcl", "COBOL", "Fortran", "CSharp", "Swift", "Kotlin",
        "GraphQL", "SPARQL", "Overpass Turbo", "Clojure"
    ]
    DATA_FORMATS = [
        "JSON", "XML", "YAML", "TOML", "CSV", "Fixlength"
    ]
    ALL_SUPPORTED = PROGRAMMING_LANGUAGES + DATA_FORMATS

    LEXER_MAP = {
        "SQL": "sql",
        "C": "c",
        "Cpp": "cpp",
        "XQuery": "xquery",
        "Java": "java",
        "Rust": "rust",
        "Erlang": "erlang",
        "Lisp": "common-lisp",
        "Bash": "bash",
        "Cmd": "doscon",
        "PowerShell": "powershell",
        "Python": "python",
        "PHP": "php",
        "CSS": "css",
        "CUDA": "cpp",
        "x86 Assembler": "nasm",
        "X86": "nasm",
        "RISC-V Assembler": "asm",
        "Riscv": "asm",
        "ARM AArch64 Assembler": "asm",
        "Arm": "asm",
        "Prolog": "prolog",
        "Java Bytecode": "jasmin",
        "Camel": "ocaml",
        "OCaml": "ocaml",
        "Go": "go",
        "Haskell": "haskell",
        "TypeScript": "typescript",
        "Tcl": "tcl",
        "COBOL": "cobol",
        "Fortran": "fortran",
        "CSharp": "csharp",
        "Swift": "swift",
        "Kotlin": "kotlin",
        "Clojure": "clojure",
        "GraphQL": "graphql",
        "SPARQL": "sparql",
        "Overpass Turbo": "text",
        "JSON": "json",
        "XML": "xml",
        "YAML": "yaml",
        "TOML": "toml",
        "CSV": "text",
        "Fixlength": "text"
    }

    def __init__(self, template_dir=None):
        if template_dir is None:
            template_dir = Path(__file__).parent / 'templates'

        self.env = Environment(
            loader=FileSystemLoader(str(template_dir)),
            autoescape=select_autoescape()
        )
        self.env.filters['format_value'] = format_value
        self.env.filters['format_table_cell'] = format_table_cell
        self.env.filters['get_lexer'] = self.get_lexer

    def _normalize(self, name: str) -> str:
        """
        Normalizes a language or format name for prefix matching by removing
        spaces, hyphens, and the word 'assembler'.
        """
        return name.lower().replace(" ", "").replace("-", "").replace("assembler", "")

    def get_lexer(self, name: str) -> str:
        """
        Determines the Pygments lexer for a given instance or language name
        using a longest-match strategy.
        """
        best_match = None
        for key in self.LEXER_MAP:
            if name.lower().startswith(self._normalize(key)) or name.lower().startswith(key.lower()):
                if best_match is None or len(key) > len(best_match):
                    best_match = key

        return self.LEXER_MAP.get(best_match, "text")

    def _get_rich_text(self, text: str, lexer_name: str) -> CellRichText:
        """
        Tokenizes the text using Pygments and returns an openpyxl CellRichText object
        with colors approximating the ReadTheDocs (RTD) style.
        """
        try:
            lexer = get_lexer_by_name(lexer_name)
        except ClassNotFound:
            return CellRichText([text])

        # Color mapping (approximating RTD 'friendly' style)
        COLOR_MAP = {
            Token.Comment: "60a0b0",
            Token.Keyword: "007020",
            Token.Keyword.Type: "902000",
            Token.Operator: "666666",
            Token.Name.Builtin: "007020",
            Token.Name.Function: "06287e",
            Token.Name.Class: "0e84b5",
            Token.Name.Namespace: "0e84b5",
            Token.Name.Exception: "007020",
            Token.Name.Variable: "bb60d5",
            Token.Name.Constant: "60add5",
            Token.Literal.String: "4070a0",
            Token.Literal.Number: "40a070",
        }

        rich_text = CellRichText()
        tokens = list(lex(text, lexer))

        for ttype, value in tokens:
            # Find the best color for the token type
            color = None
            temp_type = ttype
            while temp_type is not Token:
                if temp_type in COLOR_MAP:
                    color = COLOR_MAP[temp_type]
                    break
                temp_type = temp_type.parent

            if color:
                font = InlineFont(color=color, rFont='Consolas')
                rich_text.append(TextBlock(font, value))
            else:
                rich_text.append(TextBlock(InlineFont(rFont='Consolas'), value))

        return rich_text

    def render_pattern(self, pattern: Pattern) -> str:
        template = self.env.get_template('pattern.rst.j2')
        return template.render(pattern=pattern)

    def render_instance_table(self, pattern: Pattern, instances: List[Instance]) -> str:
        syntax_params = {"syntax"}

        display_parameters = [p for p in pattern.parameters if p.name in syntax_params]
        notes_param = next((p for p in pattern.parameters if p.name == "notes"), None)
        if notes_param:
            display_parameters.append(notes_param)

        template = self.env.get_template('instance_table.rst.j2')
        return template.render(
            pattern=pattern,
            instances=instances,
            display_parameters=display_parameters
        )

    def render_program(self, program: Program, title: str = None) -> str:
        results = []

        if title:
            results.append(f"{title}\n{'=' * len(title)}")

        # Map pattern name to pattern object
        patterns_map = {p.name: p for p in program.patterns}

        # Group instances by pattern
        instances_by_pattern: Dict[str, List[Instance]] = {}
        for instance in program.instances:
            if instance.pattern_name not in instances_by_pattern:
                instances_by_pattern[instance.pattern_name] = []
            instances_by_pattern[instance.pattern_name].append(instance)

        # Render each pattern and its instances
        for pattern_name, pattern in patterns_map.items():
            results.append(self.render_pattern(pattern))

            if pattern_name in instances_by_pattern:
                results.append(self.render_instance_table(pattern, instances_by_pattern[pattern_name]))

        return "\n\n".join(results)

    def render_pivot_table(self, program: Program, language: str) -> str:
        # Candidate parameters for columns in pivot table
        candidates = ["syntax", "notes"]

        normalized_lang = self._normalize(language)
        pivot_data = []
        for instance in program.instances:
            # First, check if any OTHER language matches the instance name better
            other_matches = [lang for lang in self.ALL_SUPPORTED if lang.lower() != language.lower()
                             and (instance.name.lower().startswith(self._normalize(lang))
                                  or instance.name.lower().startswith(lang.lower()))]

            # If the requested language matches the prefix
            # Use strict matching for short language names like 'C'
            if len(normalized_lang) <= 2:
                matches = instance.name.lower().startswith(normalized_lang) and \
                          (len(instance.name) == len(normalized_lang) or not instance.name[len(normalized_lang)].islower())
            else:
                matches = instance.name.lower().startswith(normalized_lang) or instance.name.lower().startswith(language.lower())

            if matches:
                # Ensure no OTHER language is a longer (better) match
                is_best_match = True
                for other in other_matches:
                    if len(self._normalize(other)) > len(normalized_lang):
                        is_best_match = False
                        break

                if not is_best_match:
                    continue

                assignments = {a.name: a.value for a in instance.assignments}

                pivot_data.append({
                    "pattern_name": instance.pattern_name,
                    "assignments": assignments
                })

        if not pivot_data:
            return f"No data found for language: {language}"

        # Dynamically determine which columns to show
        display_parameters = []
        for cand in candidates:
            if any(item["assignments"].get(cand, "N/A") != "N/A" for item in pivot_data):
                display_parameters.append(cand)

        # Ensure 'syntax' is at the beginning and 'notes' at the end
        if "notes" in display_parameters:
            display_parameters.remove("notes")
            display_parameters.append("notes")
        if "syntax" in display_parameters:
            display_parameters.remove("syntax")
            display_parameters.insert(0, "syntax")

        template = self.env.get_template('pivot_table.rst.j2')
        return template.render(
            language=language,
            display_parameters=display_parameters,
            pivot_data=pivot_data
        )

    def render_pivot_chapter(self, program: Program, language: str, title: str = None) -> str:
        results = []
        if not title:
            title = f"{language} Pivot View"

        results.append(f"{title}\n{'=' * len(title)}")
        results.append(self.render_pivot_table(program, language))

        return "\n\n".join(results)

    def _build_matrix_data(self, program: Program, languages: List[str], pivoted: bool = False):
        # 1. Get all pattern names in order
        pattern_names = [p.name for p in program.patterns]

        # 2. Build the matrix: row = language, column = pattern
        matrix = []

        if not pivoted:
            # Traditional matrix: Rows = Languages, Columns = Patterns
            for lang in languages:
                row = {"label": lang, "cells": []}
                for p_name in pattern_names:
                    # Find the instance for this language and pattern
                    instance = None
                    for inst in program.instances:
                        if inst.pattern_name == p_name and (inst.name.lower().startswith(self._normalize(lang)) or inst.name.lower().startswith(lang.lower())):
                            other_matches = [l for l in self.ALL_SUPPORTED if l.lower() != lang.lower()
                                             and (inst.name.lower().startswith(self._normalize(l)) or inst.name.lower().startswith(l.lower()))]
                            is_best_match = True
                            for other in other_matches:
                                if len(other) > len(lang):
                                    is_best_match = False
                                    break
                            if is_best_match:
                                instance = inst
                                break

                    if instance:
                        syntax = "N/A"
                        priority_params = ["syntax"]
                        priority_params = ["syntax"]
                        for param in priority_params:
                            val = next((a.value for a in instance.assignments if a.name == param), None)
                            if val and val != "N/A":
                                syntax = val
                                break
                    else:
                        syntax = "N/A"

                    row["cells"].append({"value": syntax, "lexer": self.get_lexer(lang)})
                matrix.append(row)
            headers = ["Language"] + pattern_names
        else:
            # Pivoted matrix: Rows = Patterns, Columns = Languages
            for p_name in pattern_names:
                row = {"label": p_name, "cells": []}
                for lang in languages:
                    instance = None
                    for inst in program.instances:
                        if inst.pattern_name == p_name and (inst.name.lower().startswith(self._normalize(lang)) or inst.name.lower().startswith(lang.lower())):
                            other_matches = [l for l in self.ALL_SUPPORTED if l.lower() != lang.lower()
                                             and (inst.name.lower().startswith(self._normalize(l)) or inst.name.lower().startswith(l.lower()))]
                            is_best_match = True
                            for other in other_matches:
                                if len(other) > len(lang):
                                    is_best_match = False
                                    break
                            if is_best_match:
                                instance = inst
                                break

                    if instance:
                        syntax = "N/A"
                        priority_params = ["syntax"]
                        for param in priority_params:
                            val = next((a.value for a in instance.assignments if a.name == param), None)
                            if val and val != "N/A":
                                syntax = val
                                break
                    else:
                        syntax = "N/A"

                    row["cells"].append({"value": syntax, "lexer": self.get_lexer(lang)})
                matrix.append(row)
            headers = ["Pattern"] + languages

        return headers, matrix

    def render_matrix_table(self, program: Program, languages: List[str], pivoted: bool = False) -> str:
        headers, matrix = self._build_matrix_data(program, languages, pivoted=pivoted)
        template = self.env.get_template('matrix_table.rst.j2')
        return template.render(
            headers=headers,
            matrix=matrix
        )

    def render_matrix_excel(self, program: Program, languages: List[str], pivoted: bool = False) -> bytes:
        headers, matrix = self._build_matrix_data(program, languages, pivoted=pivoted)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Matrix"

        # Write headers
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # Write data
        for r_idx, row in enumerate(matrix, start=2):
            # Label cell
            label_cell = ws.cell(row=r_idx, column=1, value=row["label"])
            label_cell.alignment = Alignment(vertical='top')

            for c_idx, cell_data in enumerate(row["cells"], start=2):
                value = cell_data["value"]
                lexer = cell_data["lexer"]

                cell = ws.cell(row=r_idx, column=c_idx)
                formatted_value = format_value(value)

                if formatted_value != "N/A":
                    cell.value = self._get_rich_text(formatted_value, lexer)
                else:
                    cell.value = formatted_value

                cell.alignment = Alignment(wrap_text=True, vertical='top')

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter # Get the column name
            for cell in col:
                try:
                    # For RichText, str(cell.value) gives the plain text
                    val_str = str(cell.value)
                    if val_str:
                        lines = val_str.splitlines()
                        curr_max = max(len(line) for line in lines) if lines else 0
                        if curr_max > max_length:
                            max_length = curr_max
                except:
                    pass
            # Set a reasonable max width
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    def render_matrix_chapter(self, program: Program, languages: List[str], title: str = None, pivoted: bool = False) -> str:
        results = []
        if title:
            results.append(f"{title}\n{'=' * len(title)}")

        results.append(self.render_matrix_table(program, languages, pivoted=pivoted))

        return "\n\n".join(results)

    def render_matrix_csv(self, program: Program, languages: List[str], pivoted: bool = False) -> str:
        headers, matrix = self._build_matrix_data(program, languages, pivoted=pivoted)
        template = self.env.get_template('matrix_table.csv.j2')
        return template.render(
            headers=headers,
            matrix=matrix
        )
