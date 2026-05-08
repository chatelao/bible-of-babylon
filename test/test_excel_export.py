import io
import openpyxl
from src.generator import CodeGenerator
from src.models import Program, Pattern, Instance, Parameter, Assignment

def test_render_matrix_excel_rich_text():
    # Setup dummy program
    pattern = Pattern(name="TestPattern", parameters=[Parameter(name="syntax", type="string")])
    instance = Instance(
        name="PythonInstance",
        pattern_name="TestPattern",
        assignments=[Assignment(name="syntax", value="def foo():\n    print('hello')")]
    )
    program = Program(patterns=[pattern], instances=[instance])

    generator = CodeGenerator()
    excel_data = generator.render_matrix_excel(program, languages=["Python"])

    # Load workbook from bytes
    wb = openpyxl.load_workbook(io.BytesIO(excel_data), rich_text=True)
    ws = wb.active

    # Check headers
    assert ws.cell(row=1, column=1).value == "Language"
    assert ws.cell(row=1, column=2).value == "TestPattern"
    assert ws.cell(row=1, column=1).font.bold is True

    # Check data cell
    cell = ws.cell(row=2, column=2)
    from openpyxl.cell.rich_text import CellRichText
    assert isinstance(cell.value, CellRichText)

    # Check alignment
    assert cell.alignment.wrap_text is True
    assert cell.alignment.vertical == 'top'

    # Check content (plain text part)
    assert "def foo():" in str(cell.value)
    assert "print('hello')" in str(cell.value)

if __name__ == "__main__":
    test_render_matrix_excel_rich_text()
    print("Test passed!")
